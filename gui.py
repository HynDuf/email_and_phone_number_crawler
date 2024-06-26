import tkinter as tk
from tkinter import ttk, messagebox
from threading import Thread
from queue import Queue, Empty
from crawler import parallel_crawl
from database import save_results, search_results_db
import tkinter.filedialog as filedialog
from datetime import datetime
from openpyxl import load_workbook, Workbook

class WebCrawlerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Web Crawler")
        
        self.style = ttk.Style(root)
        self.progress_var = tk.DoubleVar()
        self.setup_ui()
        self.emails = {}
        self.phone_numbers = {}
        self.queue = Queue()
        self.total_urls = 1
        self.urls_done = 0
        
    def setup_ui(self):
        root_style = ttk.Style()
        root_style.configure('.', font=('Arial', 12, 'bold'))  # Set default font size
        # add label in the layout
        self.style.layout('text.Horizontal.TProgressbar', 
                     [('Horizontal.Progressbar.trough',
                       {'children': [('Horizontal.Progressbar.pbar',
                                      {'side': 'left', 'sticky': 'ns'})],
                        'sticky': 'nswe'}), 
                      ('Horizontal.Progressbar.label', {'sticky': 'nswe'})])
        # set initial text
        self.style.configure('text.Horizontal.TProgressbar', text='0 / 0', anchor='center')
        tab_control = ttk.Notebook(self.root)
        tab_control.pack(expand=1, fill="both")
        
        crawler_tab = ttk.Frame(tab_control)
        search_tab = ttk.Frame(tab_control)
        
        tab_control.add(crawler_tab, text='Crawler')
        tab_control.add(search_tab, text='Search')
        
        self.setup_crawler_tab(crawler_tab)
        self.setup_search_tab(search_tab)
        self.setup_context_menu()

    def setup_context_menu(self):
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Copy", command=self.copy_selected_row)
        
    def treeview_right_click(self, tree, event):
        try:
            item = tree.identify_row(event.y)
            tree.focus(item)
            self.context_menu.post(event.x_root, event.y_root)
        except:
            pass

    def copy_selected_row(self):
        active_tree = self.root.focus_get()  # Get the currently focused widget
        if active_tree == self.search_results_tree:
            tree = self.search_results_tree
        elif active_tree == self.results_emails:
            tree = self.results_emails
        elif active_tree == self.results_phones:
            tree = self.results_phones
        else:
            return  # No Treeview focused, do nothing

        item = tree.selection()[0]
        values = tree.item(item, "values")
        content = " | ".join(str(value) for value in values)
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        self.show_notification("Row copied to clipboard")

    def show_notification(self, message):
        notification = tk.Toplevel(self.root)
        notification.geometry("300x50+{}+{}".format(self.root.winfo_screenwidth() - 310, self.root.winfo_screenheight() - 60))
        notification.overrideredirect(True)
        notification.attributes('-alpha', 0.9)
        notification.attributes('-topmost', True)

        label = ttk.Label(notification, text=message)
        label.pack(expand=True, fill='both', padx=10, pady=10)

        def fade():
            current_alpha = notification.attributes('-alpha')
            if current_alpha > 0:
                new_alpha = current_alpha - 0.05
                notification.attributes('-alpha', new_alpha)
                notification.after(100, fade)
            else:
                notification.destroy()

        notification.after(2000, fade)
        
    def setup_crawler_tab(self, frame):
        canvas = tk.Canvas(frame)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        scrollable_frame = ttk.Frame(canvas)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        url_label = ttk.Label(scrollable_frame, text="Start URL:")
        url_label.pack(anchor=tk.W, pady=10, padx=10)

        self.url_var = tk.StringVar()
        url_entry = ttk.Entry(scrollable_frame, textvariable=self.url_var, width=60, font=('Arial', 12))
        url_entry.pack(fill=tk.X, padx=10)

        depth_label = ttk.Label(scrollable_frame, text="Max Depth:")
        depth_label.pack(anchor=tk.W, pady=10, padx=10)

        self.depth_var = tk.IntVar(value=2)
        depth_entry = ttk.Entry(scrollable_frame, textvariable=self.depth_var, width=7, font=('Arial', 12))
        depth_entry.pack(anchor=tk.W, padx=10)

        self.start_button = ttk.Button(scrollable_frame, text="Start", command=self.start_crawl_thread, style='Large.TButton')
        self.start_button.pack(anchor=tk.W, pady=10, padx=10)

        self.links_list = tk.Listbox(scrollable_frame, width=100, height=7)
        self.links_list.pack(fill=tk.BOTH, expand=True, pady=10, padx=10)

        self.progress = ttk.Progressbar(scrollable_frame, style='text.Horizontal.TProgressbar', orient="horizontal", variable=self.progress_var)
        self.progress.pack(fill=tk.X, pady=10, padx=10)

        self.email_frame = ttk.LabelFrame(scrollable_frame, text="Email Results")
        self.email_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.results_emails = ttk.Treeview(self.email_frame, columns=("Email", "URL"), show="headings", height=6)
        self.results_emails.heading("Email", text="Email")
        self.results_emails.heading("URL", text="URL")
        self.results_emails.pack(fill=tk.BOTH, expand=True)
        self.results_emails.bind("<Button-3>", lambda event: self.treeview_right_click(self.results_emails, event))

        self.phone_frame = ttk.LabelFrame(scrollable_frame, text="Phone Results")
        self.phone_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.results_phones = ttk.Treeview(self.phone_frame, columns=("Phone Number", "URL"), show="headings", height=6)
        self.results_phones.heading("Phone Number", text="Phone Number")
        self.results_phones.heading("URL", text="URL")
        self.results_phones.pack(fill=tk.BOTH, expand=True)
        self.results_phones.bind("<Button-3>", lambda event: self.treeview_right_click(self.results_phones, event))

        button_frame = ttk.Frame(scrollable_frame)
        button_frame.pack(fill=tk.X, pady=10)

        self.save_button = ttk.Button(button_frame, text="Save results to database", command=self.save_results, style='Large.TButton')
        self.save_button.grid(row=0, column=0, padx=5, sticky=tk.W)

        self.export_button = ttk.Button(button_frame, text="Export results to Excel", command=self.export_to_excel, style='Large.TButton')
        self.export_button.grid(row=0, column=1, padx=5, sticky=tk.E)

        self.save_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)

    def setup_search_tab(self, frame):
        self.search_text = ttk.Entry(frame, width=60, font=('Arial', 14))
        self.search_text.pack(side=tk.TOP, padx=10, pady=10)
        
        search_button = ttk.Button(frame, text="Search", command=self.search_results, style='Large.TButton')
        search_button.pack(side=tk.TOP, padx=10, pady=10)

        self.search_results_tree = ttk.Treeview(frame, columns=("Type", "Content", "URL", "Root URL"), show="headings", height=10)
        self.search_results_tree.heading("Type", text="Type")
        self.search_results_tree.heading("Content", text="Content")
        self.search_results_tree.heading("URL", text="URL")
        self.search_results_tree.heading("Root URL", text="Root URL")
        self.search_results_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.search_results_tree.bind("<Button-3>", self.treeview_right_click)
        
    def start_crawl_thread(self):
        self.start_button.config(state=tk.DISABLED)
        self.save_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)
        self.links_list.delete(0, tk.END)
        self.results_emails.delete(*self.results_emails.get_children())
        self.results_phones.delete(*self.results_phones.get_children())
        self.urls_done = 0
        self.total_urls = 1
        self.progress['value'] = 0
        
        thread = Thread(target=self.start_crawl)
        thread.start()
        self.root.after(10, self.process_queue)
        
    def start_crawl(self):
        start_url = self.url_var.get()
        max_depth = self.depth_var.get()
        self.emails, self.phone_numbers = parallel_crawl(start_url, max_depth, self.queue)
        
        self.queue.put("DONE")
        
    def process_queue(self):
        try:
            if not self.queue.empty():
                msg = self.queue.get(0)
                if msg == "DONE":
                    self.start_button.config(state=tk.NORMAL)
                    self.save_button.config(state=tk.NORMAL)
                    self.export_button.config(state=tk.NORMAL)

                elif isinstance(msg, tuple):
                    if msg[0] == "URL_DONE":
                        self.urls_done += 1
                        cur_per = 100 * self.urls_done / self.total_urls
                        self.progress_var.set(cur_per)
                        self.style.configure('text.Horizontal.TProgressbar', text='{}/{}'.format(self.urls_done, self.total_urls))
                    elif msg[0] == "URL":
                        self.links_list.insert(tk.END, msg[1])
                        self.total_urls += 1
                        # Update the progress bar
                        cur_per = 100 * self.urls_done / self.total_urls
                        self.progress_var.set(cur_per)
                        self.style.configure('text.Horizontal.TProgressbar', text='{}/{}'.format(self.urls_done, self.total_urls))
                    elif msg[0] == "EMAIL":
                        self.results_emails.insert("", "end", values=(msg[1], msg[2]))
                    elif msg[0] == "PHONE":
                        self.results_phones.insert("", "end", values=(msg[1], msg[2]))
        except Empty:
            pass
        self.root.update()
        self.root.after(10, self.process_queue)
        
    def save_results(self):
        emails = {self.results_emails.item(item)['values'][0]: self.results_emails.item(item)['values'][1] for item in self.results_emails.get_children()}
        phones = {self.results_phones.item(item)['values'][0]: self.results_phones.item(item)['values'][1] for item in self.results_phones.get_children()}
        root_url = self.url_var.get()
        save_results(emails, phones, root_url)
        messagebox.showinfo("Save Results", "Results saved to database successfully!")

    def export_to_excel(self):
        # Open a file dialog to select the save location and file name
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        
        if not file_path:
            return  # User cancelled the save dialog

        # Collect metadata
        start_url = self.url_var.get()
        max_depth = self.depth_var.get()
        export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Collect URLs
        urls = [self.links_list.get(idx) for idx in range(self.links_list.size())]
        
        # Collect email results
        email_results = [(self.results_emails.item(item)['values'][0], self.results_emails.item(item)['values'][1]) for item in self.results_emails.get_children()]
        
        # Collect phone results
        phone_results = [(self.results_phones.item(item)['values'][0], self.results_phones.item(item)['values'][1]) for item in self.results_phones.get_children()]

        # Create a new workbook and select the active worksheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"

        # Write metadata
        metadata = {
            "Exported Time": export_time,
            "Root URL": start_url,
            "Max Depth": max_depth
        }
        metadata_row = 1
        for key, value in metadata.items():
            sheet.cell(row=metadata_row, column=1, value=key)
            sheet.cell(row=metadata_row, column=2, value=value)
            metadata_row += 1

        # Write headers for the results
        headers = ["URLs", "Emails", "Email Origin URLs", "Phone Numbers", "Phone Origin URLs"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=metadata_row + 1, column=col_num, value=header)
        
        # Write the actual data
        for row_num, url in enumerate(urls, metadata_row + 2):
            sheet.cell(row=row_num, column=1, value=url)
            
        for row_num, (email, url) in enumerate(email_results, metadata_row + 2):
            sheet.cell(row=row_num, column=2, value=email)
            sheet.cell(row=row_num, column=3, value=url)
        
        for row_num, (phone, url) in enumerate(phone_results, metadata_row + 2):
            sheet.cell(row=row_num, column=4, value=phone)
            sheet.cell(row=row_num, column=5, value=url)

        # Adjust column widths
        column_widths = [max(len(str(cell.value)) for cell in col) + 2 for col in sheet.columns]
        for i, column_width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = column_width

        # Save the workbook
        workbook.save(file_path)
        messagebox.showinfo("Export to Excel", "Results exported to Excel successfully!")


    
    def search_results(self):
        keyword = self.search_text.get()
        results = search_results_db(keyword)
        for row in self.search_results_tree.get_children():
            self.search_results_tree.delete(row)
        for result in results:
            self.search_results_tree.insert("", "end", values=result)
